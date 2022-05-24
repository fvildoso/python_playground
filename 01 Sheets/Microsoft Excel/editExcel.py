from openpyxl import load_workbook
import datetime


def main():
    # damos la ruta de donde editar el archivo
    nombre_archivo = 'out/mi curso favorito3.xlsx'

    # leemos el excel
    workbook = load_workbook(nombre_archivo)

    # vemos las hojas
    print(workbook.sheetnames)

    # tomamos la hoja que esté activa
    worksheet = workbook.active

    # podemos darle a una celda el valor directamente
    worksheet['B1'] = 42

    # podemos hacer un append (al final) de una fila
    worksheet.append([1, 2, 3])

    # un ejemplo con fecha
    worksheet['C1'] = datetime.datetime.now()

    # guardamos el archivo
    workbook.save(nombre_archivo)

    # lo cerramos
    workbook.close()


if __name__ == '__main__':
    main()
